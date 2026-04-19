"""
TalentCore — Excel 내보내기 유틸리티
openpyxl 기반, 공통 스타일 적용
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import make_response


# ── 공통 스타일 ────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1E3A8A")   # 딥 블루
HEADER_FONT   = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="맑은 고딕", size=10)
BOLD_FONT     = Font(name="맑은 고딕", bold=True, size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
RIGHT         = Alignment(horizontal="right",  vertical="center", wrap_text=False)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
TOTAL_FILL    = PatternFill("solid", fgColor="EFF6FF")   # 연한 파랑 (합계행)
TOTAL_FONT    = Font(name="맑은 고딕", bold=True, size=10, color="1E3A8A")
KRW_FORMAT    = '#,##0"원"'
NUM_FORMAT    = '#,##0'


def make_wb(sheet_name="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def write_header(ws, headers: list[str], row=1):
    """헤더 행 작성"""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
    ws.row_dimensions[row].height = 22


def write_row(ws, row_num: int, values: list, bold=False, total=False, align_map: dict = None):
    """데이터 행 작성"""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        if total:
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
        else:
            cell.font = BOLD_FONT if bold else BODY_FONT
        cell.border = THIN_BORDER

        # 정렬: align_map = {col_index: 'left'|'right'|'center'}
        align = (align_map or {}).get(col, 'left')
        if align == 'right':
            cell.alignment = RIGHT
        elif align == 'center':
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT


def apply_number_format(ws, col: int, start_row: int, end_row: int, fmt=KRW_FORMAT):
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=col).number_format = fmt


def auto_width(ws, min_width=8, max_width=40):
    """컬럼 너비 자동 조정"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                # 한글은 2글자 너비로 계산
                val = str(cell.value or "")
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def freeze_header(ws, row=2):
    ws.freeze_panes = ws.cell(row=row, column=1)


def to_response(wb, filename: str):
    """Flask Response로 변환"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"
    response.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return response
